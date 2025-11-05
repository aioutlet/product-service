"""
Unit tests for template generation service.
"""

import pytest
from io import BytesIO
from openpyxl import load_workbook

from src.services.template_generation_service import TemplateGenerationService


class TestTemplateGenerationService:
    """Test template generation service."""
    
    @pytest.fixture
    def service(self):
        """Create template generation service."""
        return TemplateGenerationService()
    
    def test_service_initialization(self, service):
        """Test service initializes correctly."""
        assert service is not None
        assert len(service.COMMON_COLUMNS) > 0
        assert len(service.CATEGORY_COLUMNS) > 0
    
    def test_common_columns_defined(self, service):
        """Test common columns are defined."""
        assert any(col.name == "SKU*" for col in service.COMMON_COLUMNS)
        assert any(col.name == "Product Name*" for col in service.COMMON_COLUMNS)
        assert any(col.name == "Price*" for col in service.COMMON_COLUMNS)
        assert any(col.name == "Category*" for col in service.COMMON_COLUMNS)
    
    def test_category_columns_defined(self, service):
        """Test category-specific columns are defined."""
        assert "Clothing" in service.CATEGORY_COLUMNS
        assert "Electronics" in service.CATEGORY_COLUMNS
        assert "Home & Furniture" in service.CATEGORY_COLUMNS
    
    def test_generate_clothing_template(self, service):
        """Test generating clothing category template."""
        template = service.generate_template(category="Clothing")
        
        assert template is not None
        assert isinstance(template, BytesIO)
    
    def test_generate_electronics_template(self, service):
        """Test generating electronics category template."""
        template = service.generate_template(category="Electronics")
        
        assert template is not None
        assert isinstance(template, BytesIO)
    
    def test_generate_home_template(self, service):
        """Test generating home & furniture template."""
        template = service.generate_template(category="Home & Furniture")
        
        assert template is not None
        assert isinstance(template, BytesIO)
    
    def test_template_has_data_sheet(self, service):
        """Test generated template has data sheet."""
        template = service.generate_template(category="Clothing")
        
        # Load workbook
        template.seek(0)
        wb = load_workbook(template)
        
        assert "Products" in wb.sheetnames
    
    def test_template_has_instructions_sheet(self, service):
        """Test generated template has instructions sheet."""
        template = service.generate_template(category="Clothing")
        
        # Load workbook
        template.seek(0)
        wb = load_workbook(template)
        
        assert "Instructions" in wb.sheetnames
    
    def test_template_headers_present(self, service):
        """Test template has correct headers."""
        template = service.generate_template(category="Clothing")
        
        # Load workbook
        template.seek(0)
        wb = load_workbook(template)
        ws = wb["Products"]
        
        # Check header row (row 1)
        headers = [cell.value for cell in ws[1]]
        
        assert "SKU*" in headers
        assert "Product Name*" in headers
        assert "Price*" in headers
        assert "Category*" in headers
    
    def test_template_with_examples(self, service):
        """Test template includes example rows."""
        template = service.generate_template(
            category="Clothing",
            include_examples=True
        )
        
        # Load workbook
        template.seek(0)
        wb = load_workbook(template)
        ws = wb["Products"]
        
        # Check for example data (starting row 3)
        # Row 3 should have data
        assert ws.cell(3, 1).value is not None  # SKU column
    
    def test_template_without_examples(self, service):
        """Test template without example rows."""
        template = service.generate_template(
            category="Clothing",
            include_examples=False
        )
        
        # Load workbook
        template.seek(0)
        wb = load_workbook(template)
        ws = wb["Products"]
        
        # Row 3 should be empty (no examples)
        assert ws.cell(3, 1).value is None
    
    def test_clothing_category_columns(self, service):
        """Test clothing category has specific columns."""
        template = service.generate_template(category="Clothing")
        
        # Load workbook
        template.seek(0)
        wb = load_workbook(template)
        ws = wb["Products"]
        
        headers = [cell.value for cell in ws[1]]
        
        # Check for clothing-specific columns
        assert any("Color" in str(h) for h in headers if h)
        assert any("Size" in str(h) for h in headers if h)
    
    def test_electronics_category_columns(self, service):
        """Test electronics category has specific columns."""
        template = service.generate_template(category="Electronics")
        
        # Load workbook
        template.seek(0)
        wb = load_workbook(template)
        ws = wb["Products"]
        
        headers = [cell.value for cell in ws[1]]
        
        # Check for electronics-specific columns
        assert any("Model" in str(h) for h in headers if h)
        assert any("Warranty" in str(h) for h in headers if h)
    
    def test_home_category_columns(self, service):
        """Test home & furniture category has specific columns."""
        template = service.generate_template(category="Home & Furniture")
        
        # Load workbook
        template.seek(0)
        wb = load_workbook(template)
        ws = wb["Products"]
        
        headers = [cell.value for cell in ws[1]]
        
        # Check for home-specific columns
        assert any("Dimensions" in str(h) for h in headers if h)
    
    def test_template_header_styling(self, service):
        """Test template headers have styling applied."""
        template = service.generate_template(category="Clothing")
        
        # Load workbook
        template.seek(0)
        wb = load_workbook(template)
        ws = wb["Products"]
        
        # Check first header cell has fill color (blue background)
        first_cell = ws.cell(1, 1)
        assert first_cell.fill is not None
    
    def test_invalid_category_uses_default(self, service):
        """Test invalid category uses default columns."""
        template = service.generate_template(category="InvalidCategory")
        
        assert template is not None
        assert isinstance(template, BytesIO)
