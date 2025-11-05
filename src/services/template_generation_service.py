"""
Excel Template Generation Service

Generates category-specific Excel templates for bulk product import.
"""

from typing import List, Dict, Any, Optional
from io import BytesIO
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    raise ImportError(
        "openpyxl is required for template generation. "
        "Install it with: pip install openpyxl"
    )

from src.models.bulk_import import TemplateConfig, TemplateColumn
from src.models.standard_schemas import StandardSchemas
from src.models.attribute_schema import AttributeDataType
from src.core.logger import logger


class TemplateGenerationService:
    """Service for generating Excel import templates"""
    
    # Standard column definitions for all categories
    COMMON_COLUMNS = [
        TemplateColumn(
            name="SKU*",
            field="sku",
            data_type="string",
            required=True,
            description="Unique product identifier (alphanumeric, max 100 chars)",
            max_length=100,
            example_value="TSHIRT-BLK-L-001"
        ),
        TemplateColumn(
            name="Product Name*",
            field="name",
            data_type="string",
            required=True,
            description="Product name (max 200 chars)",
            max_length=200,
            example_value="Classic Black T-Shirt - Large"
        ),
        TemplateColumn(
            name="Price*",
            field="price",
            data_type="number",
            required=True,
            description="Product price (must be positive)",
            min_value=0,
            example_value=29.99
        ),
        TemplateColumn(
            name="Description",
            field="description",
            data_type="string",
            required=False,
            description="Short product description (max 500 chars)",
            max_length=500,
            example_value="Comfortable cotton t-shirt perfect for everyday wear"
        ),
        TemplateColumn(
            name="Category*",
            field="category",
            data_type="string",
            required=True,
            description="Product category",
            example_value="Clothing"
        ),
        TemplateColumn(
            name="Brand",
            field="brand",
            data_type="string",
            required=False,
            description="Brand name",
            example_value="Urban Basics"
        ),
        TemplateColumn(
            name="Status",
            field="status",
            data_type="string",
            required=False,
            description="Product status",
            allowed_values=["active", "draft", "archived"],
            example_value="active"
        ),
        TemplateColumn(
            name="Compare At Price",
            field="compare_at_price",
            data_type="number",
            required=False,
            description="Original price before discount",
            min_value=0,
            example_value=39.99
        ),
        TemplateColumn(
            name="Tags",
            field="tags",
            data_type="string",
            required=False,
            description="Comma-separated tags",
            example_value="cotton, casual, summer"
        ),
        TemplateColumn(
            name="Image URL 1",
            field="images[0]",
            data_type="string",
            required=False,
            description="Primary product image URL",
            example_value="https://cdn.example.com/products/img1.jpg"
        ),
    ]
    
    # Category-specific additional columns
    CATEGORY_COLUMNS = {
        "Clothing": [
            TemplateColumn(
                name="Color",
                field="colors",
                data_type="string",
                required=False,
                description="Product color(s)",
                example_value="Black"
            ),
            TemplateColumn(
                name="Size",
                field="sizes",
                data_type="string",
                required=False,
                description="Available sizes",
                allowed_values=["XS", "S", "M", "L", "XL", "XXL"],
                example_value="L"
            ),
            TemplateColumn(
                name="Material",
                field="attributes.material",
                data_type="string",
                required=False,
                description="Fabric material",
                example_value="100% Cotton"
            ),
            TemplateColumn(
                name="Care Instructions",
                field="attributes.care",
                data_type="string",
                required=False,
                description="Washing/care instructions",
                example_value="Machine wash cold, tumble dry low"
            ),
        ],
        "Electronics": [
            TemplateColumn(
                name="Model Number",
                field="attributes.model",
                data_type="string",
                required=False,
                description="Manufacturer model number",
                example_value="XYZ-2024"
            ),
            TemplateColumn(
                name="Warranty (months)",
                field="attributes.warranty",
                data_type="number",
                required=False,
                description="Warranty duration in months",
                min_value=0,
                max_value=120,
                example_value=12
            ),
            TemplateColumn(
                name="Battery Required",
                field="attributes.battery_required",
                data_type="boolean",
                required=False,
                description="Does product require batteries",
                allowed_values=["Yes", "No"],
                example_value="No"
            ),
        ],
        "Home & Furniture": [
            TemplateColumn(
                name="Dimensions (L x W x H)",
                field="attributes.dimensions",
                data_type="string",
                required=False,
                description="Product dimensions in inches",
                example_value="24 x 18 x 36"
            ),
            TemplateColumn(
                name="Weight (lbs)",
                field="attributes.weight",
                data_type="number",
                required=False,
                description="Product weight in pounds",
                min_value=0,
                example_value=15.5
            ),
            TemplateColumn(
                name="Assembly Required",
                field="attributes.assembly",
                data_type="boolean",
                required=False,
                description="Requires assembly",
                allowed_values=["Yes", "No"],
                example_value="Yes"
            ),
        ],
    }
    
    def generate_template(
        self,
        category: str,
        include_examples: bool = True,
        correlation_id: Optional[str] = None
    ) -> BytesIO:
        """
        Generate Excel template for category.
        
        Args:
            category: Product category
            include_examples: Whether to include example rows
            correlation_id: Request correlation ID
            
        Returns:
            BytesIO containing Excel file
        """
        logger.info(
            f"Generating import template for category: {category}",
            metadata={"category": category, "include_examples": include_examples},
            correlation_id=correlation_id
        )
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Get columns for category
        columns = self._get_columns_for_category(category)
        
        # Add headers
        self._add_headers(ws, columns)
        
        # Add data validation
        self._add_validation(ws, columns)
        
        # Add example rows
        if include_examples:
            self._add_example_rows(ws, columns, count=3)
        
        # Add instructions sheet
        self._add_instructions_sheet(wb, category, columns)
        
        # Style the worksheet
        self._apply_styles(ws, len(columns), include_examples)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(
            f"Template generated successfully for category: {category}",
            metadata={"columns": len(columns), "size_bytes": output.getbuffer().nbytes},
            correlation_id=correlation_id
        )
        
        return output
    
    def _get_columns_for_category(self, category: str) -> List[TemplateColumn]:
        """Get column definitions for category"""
        columns = self.COMMON_COLUMNS.copy()
        
        # Add legacy category-specific columns
        if category in self.CATEGORY_COLUMNS:
            columns.extend(self.CATEGORY_COLUMNS[category])
        
        # Add structured attribute columns from StandardSchemas
        attribute_columns = self._generate_attribute_columns(category)
        columns.extend(attribute_columns)
        
        return columns
    
    def _generate_attribute_columns(self, category: str) -> List[TemplateColumn]:
        """
        Generate attribute columns from StandardSchemas for the category.
        
        Args:
            category: Product category
            
        Returns:
            List of TemplateColumn objects for category attributes
        """
        # Map category names to StandardSchemas methods
        schema_map = {
            "Clothing": StandardSchemas.get_clothing_schema,
            "Electronics": StandardSchemas.get_electronics_schema,
            "Home & Furniture": StandardSchemas.get_home_furniture_schema,
            "Beauty & Personal Care": StandardSchemas.get_beauty_schema
        }
        
        if category not in schema_map:
            return []
        
        # Get schema for category
        schema = schema_map[category]()
        
        columns = []
        
        # Process each attribute group
        for group in schema.attribute_groups:
            group_prefix = f"ATTR:{group.name}"
            
            for attr_def in group.attributes:
                # Create column name with prefix to distinguish from legacy columns
                column_name = f"{group_prefix}:{attr_def.display_name}"
                if attr_def.required:
                    column_name += "*"
                
                # Map attribute data type to template data type
                data_type_map = {
                    AttributeDataType.STRING: "string",
                    AttributeDataType.NUMBER: "number",
                    AttributeDataType.BOOLEAN: "boolean",
                    AttributeDataType.LIST: "string",  # Comma-separated
                    AttributeDataType.ENUM: "string",
                    AttributeDataType.OBJECT: "string"  # JSON string
                }
                
                template_column = TemplateColumn(
                    name=column_name,
                    field=f"structured_attributes.{group.name}.{attr_def.name}",
                    data_type=data_type_map.get(attr_def.data_type, "string"),
                    required=attr_def.required,
                    description=attr_def.description or f"{attr_def.display_name} ({attr_def.data_type.value})",
                    min_value=attr_def.min_value,
                    max_value=attr_def.max_value,
                    max_length=attr_def.max_length,
                    allowed_values=attr_def.allowed_values,
                    example_value=self._generate_example_value(attr_def)
                )
                
                columns.append(template_column)
        
        return columns
    
    def _generate_example_value(self, attr_def) -> Any:
        """Generate example value for attribute"""
        # If enum with allowed values, use first value
        if attr_def.allowed_values:
            return attr_def.allowed_values[0]
        
        # Generate based on data type
        if attr_def.data_type == AttributeDataType.STRING:
            return f"Example {attr_def.display_name.lower()}"
        elif attr_def.data_type == AttributeDataType.NUMBER:
            if attr_def.min_value is not None:
                return attr_def.min_value
            return 0
        elif attr_def.data_type == AttributeDataType.BOOLEAN:
            return "Yes"
        elif attr_def.data_type == AttributeDataType.LIST:
            return "value1, value2"
        else:
            return ""
    
    def _add_headers(self, ws, columns: List[TemplateColumn]):
        """Add header row with column names"""
        for col_idx, column in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = column.name
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Add description row
        for col_idx, column in enumerate(columns, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = column.description or ""
            cell.font = Font(italic=True, size=9)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    def _add_validation(self, ws, columns: List[TemplateColumn]):
        """Add data validation to columns"""
        for col_idx, column in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            
            # Dropdown for allowed values
            if column.allowed_values:
                dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(column.allowed_values)}"',
                    allow_blank=not column.required
                )
                dv.error = f"Please select from: {', '.join(column.allowed_values)}"
                dv.errorTitle = "Invalid Value"
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}3:{col_letter}10000")
            
            # Number validation
            elif column.data_type == "number":
                dv = DataValidation(type="decimal", allow_blank=not column.required)
                if column.min_value is not None:
                    dv.operator = "greaterThanOrEqual"
                    dv.formula1 = column.min_value
                if column.max_value is not None:
                    dv.operator = "between"
                    dv.formula1 = column.min_value or 0
                    dv.formula2 = column.max_value
                dv.error = f"Must be a number"
                if column.min_value or column.max_value:
                    dv.error += f" between {column.min_value or 'any'} and {column.max_value or 'any'}"
                dv.errorTitle = "Invalid Number"
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}3:{col_letter}10000")
    
    def _add_example_rows(self, ws, columns: List[TemplateColumn], count: int = 3):
        """Add example data rows"""
        for row_idx in range(3, 3 + count):
            for col_idx, column in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if column.example_value is not None:
                    # Modify example value slightly for each row
                    value = column.example_value
                    if isinstance(value, str) and row_idx > 3:
                        if "001" in value:
                            value = value.replace("001", f"00{row_idx - 2}")
                        elif column.field == "sku":
                            value = f"{value}-{row_idx - 2}"
                    cell.value = value
                    cell.font = Font(italic=True, color="666666")
    
    def _add_instructions_sheet(self, wb, category: str, columns: List[TemplateColumn]):
        """Add instructions worksheet"""
        ws = wb.create_sheet("Instructions")
        
        # Title
        ws.cell(row=1, column=1).value = f"Import Template Instructions - {category}"
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        # General instructions
        row = 3
        instructions = [
            "GENERAL INSTRUCTIONS:",
            "1. Fill out the 'Products' sheet with your product data",
            "2. Required fields are marked with * in the column name",
            "3. Follow the data format specified in row 2 descriptions",
            "4. Example rows are provided for reference (rows 3-5)",
            "5. You can add up to 10,000 products per file",
            "6. Save the file and upload it through the import interface",
            "",
            "IMPORT MODES:",
            "• Partial Mode: Import valid rows, skip invalid ones (recommended)",
            "• All-or-Nothing Mode: Only import if ALL rows are valid",
            "",
            "FIELD SPECIFICATIONS:",
        ]
        
        for instruction in instructions:
            ws.cell(row=row, column=1).value = instruction
            if instruction.endswith(":"):
                ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
        
        # Column details
        row += 1
        ws.cell(row=row, column=1).value = "Column Name"
        ws.cell(row=row, column=2).value = "Required"
        ws.cell(row=row, column=3).value = "Type"
        ws.cell(row=row, column=4).value = "Validation"
        for col in range(1, 5):
            ws.cell(row=row, column=col).font = Font(bold=True)
        row += 1
        
        for column in columns:
            ws.cell(row=row, column=1).value = column.name
            ws.cell(row=row, column=2).value = "Yes" if column.required else "No"
            ws.cell(row=row, column=3).value = column.data_type
            
            validation = []
            if column.allowed_values:
                validation.append(f"Values: {', '.join(column.allowed_values)}")
            if column.min_value is not None or column.max_value is not None:
                validation.append(f"Range: {column.min_value or 'any'} to {column.max_value or 'any'}")
            if column.max_length:
                validation.append(f"Max length: {column.max_length}")
            
            ws.cell(row=row, column=4).value = "; ".join(validation) if validation else "None"
            row += 1
        
        # Auto-size columns
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 30
    
    def _apply_styles(self, ws, column_count: int, has_examples: bool):
        """Apply styling to worksheet"""
        # Freeze header rows
        ws.freeze_panes = "A3"
        
        # Set column widths
        for col_idx in range(1, column_count + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Set row heights
        ws.row_dimensions[1].height = 30  # Header row
        ws.row_dimensions[2].height = 40  # Description row
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders to header and description
        for row in range(1, 3):
            for col in range(1, column_count + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        # Gray out example rows if present
        if has_examples:
            for row in range(3, 6):
                for col in range(1, column_count + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    cell.border = thin_border
