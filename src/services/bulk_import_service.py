"""
Bulk Import Service
Handles bulk product import with Excel templates, validation, and async processing.
Implements PRD REQ-5.2: Bulk Product Operations
"""
import uuid
import io
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.db.mongodb import get_product_collection
from src.models.admin_models import BulkImportJob, ImportValidationError
from src.services.dapr_publisher import get_dapr_publisher
from src.observability.logging import logger


class BulkImportService:
    """Service for bulk product import operations (REQ-5.2)"""
    
    # Template column definitions with validation rules
    TEMPLATE_COLUMNS = [
        {"field": "sku", "header": "SKU*", "required": True, 
         "description": "Unique product identifier", 
         "example": "PROD-12345"},
        {"field": "name", "header": "Product Name*", "required": True,
         "description": "Product name (max 200 chars)",
         "example": "Men's Cotton T-Shirt"},
        {"field": "description", "header": "Description", "required": False,
         "description": "Product description",
         "example": "Comfortable cotton t-shirt for everyday wear"},
        {"field": "price", "header": "Price*", "required": True,
         "description": "Price in USD (must be >= 0)",
         "example": "29.99"},
        {"field": "brand", "header": "Brand", "required": False,
         "description": "Brand name",
         "example": "Nike"},
        {"field": "department", "header": "Department", "required": False,
         "description": "Top level category (Men, Women, Kids, etc.)",
         "example": "Men"},
        {"field": "category", "header": "Category", "required": False,
         "description": "Second level category (Clothing, Shoes, etc.)",
         "example": "Clothing"},
        {"field": "subcategory", "header": "Subcategory", "required": False,
         "description": "Third level category (Tops, Bottoms, etc.)",
         "example": "Tops"},
        {"field": "tags", "header": "Tags", "required": False,
         "description": "Comma-separated tags",
         "example": "casual, summer, cotton"},
        {"field": "images", "header": "Image URLs", "required": False,
         "description": "Comma-separated image URLs",
         "example": "https://cdn.example.com/image1.jpg,https://cdn.example.com/image2.jpg"},
        {"field": "colors", "header": "Colors", "required": False,
         "description": "Comma-separated colors",
         "example": "Red, Blue, Black"},
        {"field": "sizes", "header": "Sizes", "required": False,
         "description": "Comma-separated sizes",
         "example": "S, M, L, XL"},
    ]
    
    def generate_excel_template(self, category: Optional[str] = None) -> bytes:
        """
        Generate Excel template for bulk product import (REQ-5.2.1).
        
        Args:
            category: Optional category name for category-specific template
            
        Returns:
            bytes: Excel file content
        """
        wb = Workbook()
        
        # Main data sheet
        ws = wb.active
        ws.title = "Products"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", 
                                  end_color="4472C4", 
                                  fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Add headers
        for col_idx, col_def in enumerate(self.TEMPLATE_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_def["header"]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column width
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Add example row
        for col_idx, col_def in enumerate(self.TEMPLATE_COLUMNS, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = col_def["example"]
        
        # Add instructions sheet
        instructions_ws = wb.create_sheet("Instructions")
        instructions_ws.column_dimensions['A'].width = 30
        instructions_ws.column_dimensions['B'].width = 60
        
        # Title
        title_cell = instructions_ws.cell(row=1, column=1)
        title_cell.value = "Bulk Product Import Template"
        title_cell.font = Font(bold=True, size=14)
        
        instructions_ws.cell(row=2, column=1).value = "Version:"
        instructions_ws.cell(row=2, column=2).value = "1.0"
        
        if category:
            instructions_ws.cell(row=3, column=1).value = "Category:"
            instructions_ws.cell(row=3, column=2).value = category
        
        # Field descriptions
        instructions_ws.cell(row=5, column=1).value = "Field"
        instructions_ws.cell(row=5, column=1).font = Font(bold=True)
        instructions_ws.cell(row=5, column=2).value = "Description"
        instructions_ws.cell(row=5, column=2).font = Font(bold=True)
        
        row = 6
        for col_def in self.TEMPLATE_COLUMNS:
            instructions_ws.cell(row=row, column=1).value = col_def["header"]
            instructions_ws.cell(row=row, column=2).value = col_def["description"]
            row += 1
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def validate_import_data(
        self, 
        file_content: bytes, 
        filename: str
    ) -> tuple[List[Dict[str, Any]], List[ImportValidationError]]:
        """
        Validate Excel import file and extract product data (REQ-5.2.2).
        
        Args:
            file_content: Excel file content as bytes
            filename: Original filename
            
        Returns:
            Tuple of (valid_products, validation_errors)
        """
        errors = []
        valid_products = []
        
        try:
            # Load workbook
            wb = load_workbook(io.BytesIO(file_content))
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col_idx).value
                if cell_value:
                    headers.append(str(cell_value).strip())
            
            # Process data rows (skip header)
            for row_idx in range(2, ws.max_row + 1):
                row_data = {}
                has_data = False
                
                # Extract cell values
                for col_idx, header in enumerate(headers, start=1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        has_data = True
                        row_data[header] = str(cell_value).strip()
                
                # Skip empty rows
                if not has_data:
                    continue
                
                # Map headers to field names and validate
                product_data = {}
                row_errors = []
                
                for col_def in self.TEMPLATE_COLUMNS:
                    header = col_def["header"]
                    field = col_def["field"]
                    value = row_data.get(header, "").strip()
                    
                    # Check required fields
                    if col_def["required"] and not value:
                        row_errors.append(
                            ImportValidationError(
                                row_number=row_idx,
                                field_name=field,
                                error_description=f"{header} is required",
                                suggested_correction=f"Provide a value for {header}",
                                current_value=value or "empty"
                            )
                        )
                        continue
                    
                    # Field-specific validation
                    if field == "price" and value:
                        try:
                            price = float(value)
                            if price < 0:
                                row_errors.append(
                                    ImportValidationError(
                                        row_number=row_idx,
                                        field_name=field,
                                        error_description="Price must be non-negative",
                                        suggested_correction="Provide a price >= 0",
                                        current_value=value
                                    )
                                )
                            else:
                                product_data[field] = price
                        except ValueError:
                            row_errors.append(
                                ImportValidationError(
                                    row_number=row_idx,
                                    field_name=field,
                                    error_description="Price must be a valid number",
                                    suggested_correction="Provide a numeric value",
                                    current_value=value
                                )
                            )
                    elif field in ["tags", "images", "colors", "sizes"] and value:
                        # Convert comma-separated to list
                        product_data[field] = [
                            item.strip() for item in value.split(",") if item.strip()
                        ]
                    elif value:
                        product_data[field] = value
                
                # If row has errors, add to errors list
                if row_errors:
                    errors.extend(row_errors)
                else:
                    # Add required metadata
                    product_data["created_by"] = "bulk_import"
                    product_data["is_active"] = True
                    valid_products.append(product_data)
            
            logger.info(
                f"Validated import file: {filename}",
                metadata={
                    'event': 'import_validation_complete',
                    'validProducts': len(valid_products),
                    'errorCount': len(errors)
                }
            )
            
        except Exception as e:
            logger.error(
                f"Failed to parse import file: {str(e)}",
                metadata={'error': str(e), 'filename': filename}
            )
            errors.append(
                ImportValidationError(
                    row_number=0,
                    field_name="file",
                    error_description=f"Failed to parse file: {str(e)}",
                    suggested_correction="Ensure file is valid Excel format",
                    current_value=filename
                )
            )
        
        return valid_products, errors
    
    async def create_import_job(
        self,
        filename: str,
        total_rows: int,
        created_by: str,
        import_mode: str = "partial"
    ) -> BulkImportJob:
        """
        Create a new bulk import job record (REQ-5.2.4).
        
        Args:
            filename: Original filename
            total_rows: Total rows to import
            created_by: User ID creating the job
            import_mode: "partial" or "all-or-nothing"
            
        Returns:
            BulkImportJob: Created job record
        """
        job_id = str(uuid.uuid4())
        
        job = BulkImportJob(
            job_id=job_id,
            filename=filename,
            status="pending",
            total_rows=total_rows,
            created_by=created_by,
            import_mode=import_mode
        )
        
        # Store job in database
        collection = await get_product_collection()
        db = collection.database
        jobs_collection = db.get_collection("import_jobs")
        
        await jobs_collection.insert_one(job.dict())
        
        logger.info(
            f"Created import job: {job_id}",
            metadata={
                'event': 'import_job_created',
                'jobId': job_id,
                'totalRows': total_rows,
                'createdBy': created_by
            }
        )
        
        return job
    
    async def update_job_status(
        self,
        job_id: str,
        status: str,
        processed_rows: int = None,
        success_count: int = None,
        error_count: int = None,
        error_report_url: str = None
    ):
        """
        Update import job status (REQ-5.2.4).
        
        Args:
            job_id: Job ID
            status: New status (processing, completed, failed, cancelled)
            processed_rows: Number of rows processed
            success_count: Number of successful imports
            error_count: Number of errors
            error_report_url: URL to error report
        """
        collection = await get_product_collection()
        db = collection.database
        jobs_collection = db.get_collection("import_jobs")
        
        update_data = {"status": status}
        
        if processed_rows is not None:
            update_data["processed_rows"] = processed_rows
        if success_count is not None:
            update_data["success_count"] = success_count
        if error_count is not None:
            update_data["error_count"] = error_count
        if error_report_url:
            update_data["error_report_url"] = error_report_url
        
        if status in ["completed", "failed", "cancelled"]:
            update_data["completed_at"] = datetime.now(timezone.utc)
        
        await jobs_collection.update_one(
            {"job_id": job_id},
            {"$set": update_data}
        )
        
        logger.info(
            f"Updated job status: {job_id} -> {status}",
            metadata={
                'event': 'import_job_updated',
                'jobId': job_id,
                'status': status
            }
        )
    
    async def get_import_job(self, job_id: str) -> Optional[Dict[str, Any]]:
        """Get import job by ID (REQ-5.2.4)."""
        collection = await get_product_collection()
        db = collection.database
        jobs_collection = db.get_collection("import_jobs")
        
        job = await jobs_collection.find_one({"job_id": job_id})
        return job
    
    async def list_import_jobs(
        self,
        skip: int = 0,
        limit: int = 20,
        status: Optional[str] = None
    ) -> tuple[List[Dict[str, Any]], int]:
        """
        List import jobs with pagination (REQ-5.2.4).
        
        Returns:
            Tuple of (jobs list, total count)
        """
        collection = await get_product_collection()
        db = collection.database
        jobs_collection = db.get_collection("import_jobs")
        
        query = {}
        if status:
            query["status"] = status
        
        cursor = jobs_collection.find(query).sort(
            "started_at", -1
        ).skip(skip).limit(limit)
        
        jobs = await cursor.to_list(length=limit)
        total = await jobs_collection.count_documents(query)
        
        return jobs, total
    
    async def publish_import_job_event(
        self,
        job_id: str,
        event_type: str,
        data: Dict[str, Any]
    ):
        """
        Publish import job events (REQ-5.2.4).
        
        Event types: product.bulk.import.job.created,
                    product.bulk.import.progress,
                    product.bulk.import.completed,
                    product.bulk.import.failed
        """
        publisher = get_dapr_publisher()
        await publisher.publish(
            event_type,
            {
                "jobId": job_id,
                **data
            },
            None
        )


# Singleton instance
_bulk_import_service = None


def get_bulk_import_service() -> BulkImportService:
    """Get singleton bulk import service instance."""
    global _bulk_import_service
    if _bulk_import_service is None:
        _bulk_import_service = BulkImportService()
    return _bulk_import_service
