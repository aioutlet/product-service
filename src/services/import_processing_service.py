"""
Import Processing Service

Handles Excel file parsing, validation, and product import.
"""

from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, timezone
from io import BytesIO
import uuid
import re

try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError(
        "openpyxl is required for import processing. "
        "Install it with: pip install openpyxl"
    )

from src.models.bulk_import import (
    ImportJob,
    ImportStatus,
    ImportMode,
    ImportJobProgress,
    ValidationError,
    ValidationSeverity
)
from src.models.product import ProductCreate
from src.models.attribute_schema import ProductAttributes
from src.services.attribute_validation_service import AttributeValidationService
from src.repositories.product_repository import ProductRepository
from src.services.dapr_publisher import get_dapr_publisher
from src.core.logger import logger
from src.core.errors import ErrorResponse


class ImportProcessingService:
    """Service for processing product imports from Excel"""
    
    def __init__(self, repository: ProductRepository):
        """
        Initialize import processing service.
        
        Args:
            repository: Product repository for data access
        """
        self.repository = repository
        self.max_file_size = 50 * 1024 * 1024  # 50MB
        self.max_products = 10000
    
    async def create_import_job(
        self,
        file_content: bytes,
        filename: str,
        mode: ImportMode,
        created_by: str,
        category: Optional[str] = None,
        correlation_id: Optional[str] = None
    ) -> ImportJob:
        """
        Create a new import job.
        
        Args:
            file_content: Excel file content
            filename: Original filename
            mode: Import mode (partial or all-or-nothing)
            created_by: User ID
            category: Product category
            correlation_id: Request correlation ID
            
        Returns:
            ImportJob record
            
        Raises:
            ErrorResponse: If file is invalid
        """
        # Validate file size
        file_size = len(file_content)
        if file_size > self.max_file_size:
            raise ErrorResponse(
                f"File size {file_size} bytes exceeds maximum {self.max_file_size} bytes",
                status_code=400
            )
        
        # Validate file extension
        if not filename.lower().endswith(('.xlsx', '.xls')):
            raise ErrorResponse(
                "Invalid file format. Only .xlsx and .xls files are supported.",
                status_code=400
            )
        
        # Create job record
        job = ImportJob(
            job_id=str(uuid.uuid4()),
            filename=filename,
            file_size=file_size,
            mode=mode,
            status=ImportStatus.PENDING,
            category=category,
            created_by=created_by,
            created_at=datetime.now(timezone.utc),
            correlation_id=correlation_id,
            progress=ImportJobProgress(total_rows=0)
        )
        
        logger.info(
            f"Created import job {job.job_id}",
            metadata={
                "filename": filename,
                "size": file_size,
                "mode": mode,
                "created_by": created_by
            },
            correlation_id=correlation_id
        )
        
        return job
    
    async def validate_import(
        self,
        job: ImportJob,
        file_content: bytes,
        correlation_id: Optional[str] = None
    ) -> Tuple[List[Dict[str, Any]], List[ValidationError], List[ValidationError]]:
        """
        Validate Excel file and extract product data.
        
        Args:
            job: Import job record
            file_content: Excel file content
            correlation_id: Request correlation ID
            
        Returns:
            Tuple of (parsed_products, errors, warnings)
        """
        logger.info(
            f"Validating import file for job {job.job_id}",
            correlation_id=correlation_id
        )
        
        job.status = ImportStatus.VALIDATING
        
        errors: List[ValidationError] = []
        warnings: List[ValidationError] = []
        products: List[Dict[str, Any]] = []
        
        try:
            # Load workbook
            wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
            ws = wb.active
            
            # Get headers from row 1
            headers = [cell.value for cell in ws[1]]
            
            # Update total rows
            total_rows = ws.max_row - 2  # Exclude header and description rows
            job.progress.total_rows = total_rows
            
            if total_rows == 0:
                errors.append(ValidationError(
                    row_number=0,
                    error_code="EMPTY_FILE",
                    error_message="No data rows found in file"
                ))
                return products, errors, warnings
            
            if total_rows > self.max_products:
                errors.append(ValidationError(
                    row_number=0,
                    error_code="TOO_MANY_ROWS",
                    error_message=f"File contains {total_rows} rows, maximum is {self.max_products}"
                ))
                return products, errors, warnings
            
            # Track SKUs for duplicate detection
            seen_skus = set()
            
            # Parse data rows (starting from row 3)
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                # Skip empty rows
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                
                # Parse row
                product_data = {}
                structured_attrs = {}  # For structured_attributes
                row_errors = []
                row_warnings = []
                
                for col_idx, (header, value) in enumerate(zip(headers, row), start=1):
                    if not header:
                        continue
                    
                    # Clean header (remove * for required fields)
                    field_name = header.replace('*', '').strip()
                    
                    # Check if this is a structured attribute column (starts with ATTR:)
                    if field_name.startswith('ATTR:'):
                        # Parse structured attribute: "ATTR:group_name:Display Name"
                        parts = field_name.split(':', 2)
                        if len(parts) == 3:
                            _, group_name, attr_display = parts
                            
                            # Convert group name and attribute display to field names
                            # e.g., "ATTR:Materials & Composition:Primary Material" -> structured_attrs['materials_composition']['primary_material']
                            group_field = group_name.lower().replace(' & ', '_').replace(' ', '_')
                            attr_field = attr_display.lower().replace(' ', '_')
                            
                            if group_field not in structured_attrs:
                                structured_attrs[group_field] = {}
                            
                            # Parse value based on type
                            if value is not None and str(value).strip():
                                # Handle boolean values
                                if str(value).strip().lower() in ['yes', 'no', 'true', 'false']:
                                    structured_attrs[group_field][attr_field] = str(value).strip().lower() in ['yes', 'true']
                                # Handle numeric values
                                elif isinstance(value, (int, float)):
                                    structured_attrs[group_field][attr_field] = value
                                # Handle lists (comma-separated)
                                elif ',' in str(value):
                                    structured_attrs[group_field][attr_field] = [v.strip() for v in str(value).split(',')]
                                else:
                                    structured_attrs[group_field][attr_field] = str(value).strip()
                        continue
                    
                    # Map header to field for regular columns
                    field = self._map_header_to_field(field_name)
                    
                    # Validate and convert value
                    validated_value, field_errors = self._validate_field(
                        field=field,
                        value=value,
                        row_number=row_idx,
                        required='*' in header
                    )
                    
                    if field_errors:
                        row_errors.extend(field_errors)
                    elif validated_value is not None:
                        # Handle nested fields (e.g., "attributes.material")
                        if '.' in field:
                            parts = field.split('.')
                            if parts[0] not in product_data:
                                product_data[parts[0]] = {}
                            product_data[parts[0]][parts[1]] = validated_value
                        else:
                            product_data[field] = validated_value
                
                # Add structured attributes if any were found
                if structured_attrs:
                    # Convert to ProductAttributes and validate
                    try:
                        attrs_obj = ProductAttributes(**structured_attrs)
                        product_data['structured_attributes'] = attrs_obj.model_dump()
                        
                        # Validate attributes if category is known
                        if 'category' in product_data:
                            validation_service = AttributeValidationService()
                            validation_result = validation_service.validate_attributes(
                                attrs_obj,
                                product_data['category'],
                                correlation_id=correlation_id
                            )
                            
                            if not validation_result.is_valid:
                                for error in validation_result.errors:
                                    row_errors.append(ValidationError(
                                        row_number=row_idx,
                                        field_name=error.field_path,
                                        error_code="INVALID_ATTRIBUTE",
                                        error_message=error.message,
                                        current_value=str(error.invalid_value) if error.invalid_value else None
                                    ))
                    except Exception as e:
                        row_warnings.append(ValidationError(
                            row_number=row_idx,
                            field_name="structured_attributes",
                            error_code="ATTRIBUTE_PARSE_ERROR",
                            error_message=f"Failed to parse attributes: {str(e)}",
                            severity=ValidationSeverity.WARNING
                        ))
                
                # Check for duplicate SKU in file
                sku = product_data.get('sku')
                if sku:
                    if sku in seen_skus:
                        row_errors.append(ValidationError(
                            row_number=row_idx,
                            field_name="SKU",
                            error_code="DUPLICATE_SKU",
                            error_message=f"Duplicate SKU '{sku}' found in file",
                            current_value=sku
                        ))
                    seen_skus.add(sku)
                
                # Add validation results
                errors.extend(row_errors)
                warnings.extend(row_warnings)
                
                # Add product if no errors (or in partial mode)
                if not row_errors:
                    products.append(product_data)
            
            # Check for duplicate SKUs in database
            if products:
                skus = [p['sku'] for p in products if 'sku' in p]
                existing = await self.repository.collection.find(
                    {"sku": {"$in": skus}, "is_active": True}
                ).to_list(length=None)
                
                if existing:
                    for existing_product in existing:
                        errors.append(ValidationError(
                            row_number=0,  # Can't determine row number easily
                            field_name="SKU",
                            error_code="SKU_EXISTS",
                            error_message=f"SKU '{existing_product['sku']}' already exists in database",
                            current_value=existing_product['sku']
                        ))
            
            logger.info(
                f"Validation completed for job {job.job_id}",
                metadata={
                    "products": len(products),
                    "errors": len(errors),
                    "warnings": len(warnings)
                },
                correlation_id=correlation_id
            )
            
        except Exception as e:
            logger.error(
                f"Failed to validate import file: {str(e)}",
                correlation_id=correlation_id
            )
            errors.append(ValidationError(
                row_number=0,
                error_code="FILE_PARSE_ERROR",
                error_message=f"Failed to parse Excel file: {str(e)}"
            ))
        
        return products, errors, warnings
    
    async def process_import(
        self,
        job: ImportJob,
        products: List[Dict[str, Any]],
        created_by: str,
        correlation_id: Optional[str] = None
    ) -> ImportJob:
        """
        Process the import and create products.
        
        Args:
            job: Import job record
            products: Validated product data
            created_by: User ID
            correlation_id: Request correlation ID
            
        Returns:
            Updated ImportJob
        """
        logger.info(
            f"Processing import for job {job.job_id}",
            metadata={"product_count": len(products)},
            correlation_id=correlation_id
        )
        
        job.status = ImportStatus.PROCESSING
        job.started_at = datetime.now(timezone.utc)
        job.progress.processed_rows = 0
        job.progress.successful_rows = 0
        job.progress.failed_rows = 0
        
        imported_ids = []
        
        for idx, product_data in enumerate(products):
            try:
                # Add required fields
                product_data['created_by'] = created_by
                product_data['created_at'] = datetime.now(timezone.utc)
                product_data['updated_at'] = datetime.now(timezone.utc)
                product_data['is_active'] = True
                product_data['history'] = []
                
                # Insert product
                result = await self.repository.collection.insert_one(product_data)
                imported_ids.append(str(result.inserted_id))
                
                job.progress.successful_rows += 1
                
                # Publish product.created event
                try:
                    publisher = get_dapr_publisher()
                    await publisher.publish_product_created(
                        product_id=str(result.inserted_id),
                        product_data=product_data,
                        created_by=created_by,
                        correlation_id=correlation_id
                    )
                except Exception as e:
                    logger.warning(f"Failed to publish product.created event: {str(e)}")
                
            except Exception as e:
                logger.warning(
                    f"Failed to import product at index {idx}: {str(e)}",
                    correlation_id=correlation_id
                )
                job.progress.failed_rows += 1
            
            job.progress.processed_rows += 1
            job.progress.current_row = idx + 1
            job.progress.percentage = (job.progress.processed_rows / len(products)) * 100
        
        # Update job status
        job.imported_product_ids = imported_ids
        job.completed_at = datetime.now(timezone.utc)
        
        if job.progress.failed_rows == 0:
            job.status = ImportStatus.COMPLETED
            job.summary_message = f"Successfully imported {job.progress.successful_rows} products"
        elif job.progress.successful_rows > 0:
            job.status = ImportStatus.PARTIAL
            job.summary_message = (
                f"Imported {job.progress.successful_rows} products, "
                f"{job.progress.failed_rows} failed"
            )
        else:
            job.status = ImportStatus.FAILED
            job.summary_message = f"Import failed: {job.progress.failed_rows} products could not be imported"
        
        # Publish bulk.completed event
        try:
            publisher = get_dapr_publisher()
            await publisher.publish_bulk_operation_completed(
                operation="import",
                success_count=job.progress.successful_rows,
                failure_count=job.progress.failed_rows,
                total_count=len(products),
                executed_by=created_by,
                details={"job_id": job.job_id, "filename": job.filename},
                correlation_id=correlation_id
            )
        except Exception as e:
            logger.error(f"Failed to publish bulk.completed event: {str(e)}")
        
        logger.info(
            f"Import completed for job {job.job_id}",
            metadata={
                "status": job.status,
                "successful": job.progress.successful_rows,
                "failed": job.progress.failed_rows
            },
            correlation_id=correlation_id
        )
        
        return job
    
    def _map_header_to_field(self, header: str) -> str:
        """Map Excel header to product field"""
        mapping = {
            "SKU": "sku",
            "Product Name": "name",
            "Price": "price",
            "Description": "description",
            "Category": "category",
            "Brand": "brand",
            "Status": "status",
            "Compare At Price": "compare_at_price",
            "Tags": "tags",
            "Image URL 1": "images",
            "Color": "colors",
            "Size": "sizes",
            "Material": "attributes.material",
            "Care Instructions": "attributes.care",
            "Model Number": "attributes.model",
            "Warranty (months)": "attributes.warranty",
            "Battery Required": "attributes.battery_required",
            "Dimensions (L x W x H)": "attributes.dimensions",
            "Weight (lbs)": "attributes.weight",
            "Assembly Required": "attributes.assembly",
        }
        return mapping.get(header, header.lower().replace(' ', '_'))
    
    def _validate_field(
        self,
        field: str,
        value: Any,
        row_number: int,
        required: bool = False
    ) -> Tuple[Optional[Any], List[ValidationError]]:
        """
        Validate a field value.
        
        Returns:
            Tuple of (validated_value, errors)
        """
        errors = []
        
        # Check if empty
        if value is None or (isinstance(value, str) and value.strip() == ''):
            if required:
                errors.append(ValidationError(
                    row_number=row_number,
                    field_name=field,
                    error_code="REQUIRED_FIELD_MISSING",
                    error_message=f"Required field '{field}' is missing",
                    severity=ValidationSeverity.ERROR
                ))
            return None, errors
        
        # Convert and validate based on field
        try:
            if field == "price" or field == "compare_at_price":
                val = float(value)
                if val < 0:
                    errors.append(ValidationError(
                        row_number=row_number,
                        field_name=field,
                        error_code="INVALID_PRICE",
                        error_message=f"Price must be non-negative, got {val}",
                        current_value=value
                    ))
                return val, errors
            
            elif field == "sku":
                val = str(value).strip()
                if len(val) > 100:
                    errors.append(ValidationError(
                        row_number=row_number,
                        field_name=field,
                        error_code="SKU_TOO_LONG",
                        error_message=f"SKU exceeds 100 characters",
                        current_value=val
                    ))
                if not re.match(r'^[A-Za-z0-9\-_]+$', val):
                    errors.append(ValidationError(
                        row_number=row_number,
                        field_name=field,
                        error_code="INVALID_SKU_FORMAT",
                        error_message=f"SKU contains invalid characters (use only alphanumeric, dash, underscore)",
                        current_value=val,
                        suggested_correction="Use only letters, numbers, dashes, and underscores"
                    ))
                return val, errors
            
            elif field == "status":
                val = str(value).strip().lower()
                if val not in ["active", "draft", "archived"]:
                    errors.append(ValidationError(
                        row_number=row_number,
                        field_name=field,
                        error_code="INVALID_STATUS",
                        error_message=f"Status must be 'active', 'draft', or 'archived'",
                        current_value=value,
                        suggested_correction="Use 'active', 'draft', or 'archived'"
                    ))
                return val, errors
            
            elif field == "tags":
                # Split comma-separated tags
                val = [tag.strip() for tag in str(value).split(',') if tag.strip()]
                return val, errors
            
            elif field == "images":
                # Single image URL
                val = str(value).strip()
                if val and not val.startswith(('http://', 'https://')):
                    errors.append(ValidationError(
                        row_number=row_number,
                        field_name=field,
                        error_code="INVALID_IMAGE_URL",
                        error_message=f"Image URL must start with http:// or https://",
                        current_value=val
                    ))
                return [val] if val else [], errors
            
            else:
                # Default: convert to string
                return str(value).strip() if value else None, errors
            
        except (ValueError, TypeError) as e:
            errors.append(ValidationError(
                row_number=row_number,
                field_name=field,
                error_code="TYPE_CONVERSION_ERROR",
                error_message=f"Failed to convert value: {str(e)}",
                current_value=value
            ))
            return None, errors
